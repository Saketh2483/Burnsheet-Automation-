#!/usr/bin/env python
import openpyxl
from openpyxl.utils import get_column_letter
import os
from pathlib import Path
from datetime import datetime

def combine_excel_sheets(input_file, output_file):
    """
    Combines multiple Excel sheets into a master data file with specified logic
    """
    
    # Load the input workbook - use data_only=True to get calculated values instead of formulas
    wb_input = openpyxl.load_workbook(input_file, data_only=True)
    
    # Get all sheet names, skip hidden sheets and sheets named "legend"
    sheet_names = []
    for sheet_name in wb_input.sheetnames:
        ws = wb_input[sheet_name]
        # Skip hidden sheets (sheet_state will be 'hidden')
        if ws.sheet_state != 'hidden':
            # Skip sheets named "legend" (case-insensitive)
            if sheet_name.lower() != 'legend':
                sheet_names.append(sheet_name)
    
    print(f"Processing sheets: {sheet_names}")
    
    # Create new output workbook
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = "Combined Data"
    
    # Define headers in the correct order as per specification
    headers = [
        "EMPId", "Name", "Location", "Country", "ACT/PCT", "Skill Set",
        "Verizon Level Mapping", "Classification", "Key", "Cognizant Designation",
        "ESA ID", "ESA Description", "Service Line", "Timesheet",
        "Hourly Rate(Rs)", "Hourly Rate($)", "Projected Rate($)", "Actual Rate", "Variance",
        "Jan-26", "Feb-26", "Mar-26", "Verizon TQ ID", "Verizon TQ Description", "POC"
    ]
    
    # Write headers only once
    for col_idx, header in enumerate(headers, 1):
        ws_output.cell(row=1, column=col_idx, value=header)
    
    # Header row is always row 8 as per requirement
    header_row_number = 8
    print(f"Header row number: {header_row_number}")
    
    # Process each sheet
    current_row = 2
    india_sheets_count = 0
    usa_sheets_count = 0
    
    for sheet_name in sheet_names:
        ws_input = wb_input[sheet_name]
        
        # Variable x stores the sheet name
        x = sheet_name
        print(f"Sheet name = {x}")
        
        # Determine country based on sheet name
        if x and x[0].isdigit():
            country = "India"
            india_sheets_count += 1
        elif x and x[0].isalpha():
            country = "USA"
            usa_sheets_count += 1
        else:
            continue
        
        # Read VerizonTqId from row 4, columns C:E (combine them)
        verizon_tq_id = ""
        for col in ['C', 'D', 'E']:
            cell_value = ws_input[f'{col}4'].value
            if cell_value:
                verizon_tq_id += str(cell_value)
        
        # Read VerizonTqDescription from row 5, columns C:E (combine them)
        verizon_tq_description = ""
        for col in ['C', 'D', 'E']:
            cell_value = ws_input[f'{col}5'].value
            if cell_value:
                verizon_tq_description += str(cell_value)
        
        # Read header row from the consistently determined header_row_number
        header_row_mapping = {}
        for col in range(1, ws_input.max_column + 1):
            header_value = ws_input.cell(row=header_row_number, column=col).value
            if header_value:
                # Keep datetime objects as-is, convert others to string
                if isinstance(header_value, datetime):
                    header_row_mapping[col] = header_value
                else:
                    # Only map if this header hasn't been mapped yet (avoid duplicates)
                    header_str = str(header_value).strip()
                    # Check if this header name already exists in our mapping
                    header_exists = any(v == header_str for v in header_row_mapping.values())
                    if not header_exists:
                        header_row_mapping[col] = header_str
        
        # Process data rows (starting from the row after the header row)
        for row in range(header_row_number + 1, ws_input.max_row + 1):
            # Skip hidden rows
            if ws_input.row_dimensions[row].hidden:
                continue
            
            # Check if Name column has value (skip if empty)
            name_value = None
            name_col_idx = None
            for col, header in header_row_mapping.items():
                if 'name' in header.lower():
                    name_value = ws_input.cell(row=row, column=col).value
                    name_col_idx = col
                    break
            
            if not name_value:
                continue
            
            # Build row data from input, reading all columns
            row_data = {}
            
            for col, header in header_row_mapping.items():
                if header:
                    cell_value = ws_input.cell(row=row, column=col).value
                    # Only add if we haven't seen this header before (first occurrence wins)
                    if header not in row_data:
                        row_data[header] = cell_value
            
            # Process and map data to output columns
            output_row = [None] * len(headers)
            
            # Map basic columns from input data
            for header_idx, header in enumerate(headers):
                if header == "EMPId":
                    # Look for EmpID or EMPId or similar in row_data
                    emp_value = None
                    if "EmpID" in row_data:
                        emp_value = row_data["EmpID"]
                    elif "EMPId" in row_data:
                        emp_value = row_data["EMPId"]
                    else:
                        # Search for any empid-like key
                        for key, value in row_data.items():
                            if key and isinstance(key, str) and 'empid' in key.lower():
                                emp_value = value
                                break
                    if emp_value is not None:
                        output_row[header_idx] = emp_value
                elif header == "Country":
                    output_row[header_idx] = country
                elif header == "Verizon TQ ID":
                    output_row[header_idx] = verizon_tq_id
                elif header == "Verizon TQ Description":
                    output_row[header_idx] = verizon_tq_description
                elif header == "POC":
                    output_row[header_idx] = x
                elif header in ["Jan-26", "Feb-26", "Mar-26"]:
                    # Fill month columns with Projected Rate($) value
                    projected_rate = row_data.get("Projected Rate($)", None)
                    if projected_rate is not None:
                        output_row[header_idx] = projected_rate
                else:
                    # Try to find the matching column in row_data with flexible matching
                    value_found = False
                    
                    # First try exact match
                    if header in row_data:
                        output_row[header_idx] = row_data[header]
                        value_found = True
                    else:
                        # Try case-insensitive match with flexible whitespace
                        for key, value in row_data.items():
                            if key and isinstance(key, str):
                                if key.strip().lower() == header.strip().lower():
                                    output_row[header_idx] = value
                                    value_found = True
                                    break
                        
                        # Special handling for ESA Description - try alternative names
                        if not value_found and header == "ESA Description":
                            for key, value in row_data.items():
                                if key and isinstance(key, str):
                                    key_lower = key.strip().lower()
                                    if 'esa' in key_lower and 'description' in key_lower:
                                        output_row[header_idx] = value
                                        value_found = True
                                        break
                            
                            # If still not found, try "Description" alone
                            if not value_found:
                                for key, value in row_data.items():
                                    if key and isinstance(key, str):
                                        if key.strip().lower() == 'description':
                                            output_row[header_idx] = value
                                            value_found = True
                                            break
            
            # Handle Classification and Key based on country
            classification = None
            verizon_level = row_data.get("Verizon Level Mapping", "")
            location = row_data.get("Location", "")
            
            if country == "India":
                # Classification: Read directly from Classification column
                classification = row_data.get("Classification", "")
                output_row[headers.index("Classification")] = classification
                # Key = Classification + Location + Verizon Level Mapping (no delimiters)
                key_parts = [str(p) for p in [classification, location, verizon_level] if p]
                key = "".join(key_parts)
                output_row[headers.index("Key")] = key
            else:  # USA
                # Classification = Verizon Level Mapping with characters before and including '-' removed, then remove "Skills"
                if verizon_level:
                    verizon_level_str = str(verizon_level)
                    if '-' in verizon_level_str:
                        # Remove all characters up to and including '-'
                        classification = verizon_level_str.split('-', 1)[1]
                    else:
                        classification = verizon_level_str
                    # Remove the word "Skills" from the classification
                    classification = classification.replace("Skills", "").strip()
                
                output_row[headers.index("Classification")] = classification
                
                # Key = Location + Verizon Level Mapping (no delimiters)
                key_parts = [str(p) for p in [location, verizon_level] if p]
                key = "".join(key_parts)
                output_row[headers.index("Key")] = key
            
            # Write row to output
            for col_idx, value in enumerate(output_row, 1):
                ws_output.cell(row=current_row, column=col_idx, value=value)
            
            current_row += 1
    
    # Save output workbook
    wb_output.save(output_file)
    print(f"\n--- Processing Summary ---")
    print(f"India Sheets Processed: {india_sheets_count}")
    print(f"USA Sheets Processed: {usa_sheets_count}")
    print(f"Total Rows Combined: {current_row - 2}")
    print(f"\nCombined data saved to {output_file}")

if __name__ == "__main__":
    input_file = "public/Home & Marketing, SOR - Burn 2026.xlsx"
    output_file = "public/Combined-Input.xlsx"
    
    if os.path.exists(input_file):
        combine_excel_sheets(input_file, output_file)
    else:
        print(f"Error: {input_file} not found")
