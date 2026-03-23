"""
Rate Calculation Script for Burnsheet Automation
Reads rate cards from Excel files and populates rate columns in Combined-Input.xlsx
"""

from openpyxl import load_workbook
import os

# File paths
INDIA_RATE_FILE = r"C:\Verizon Project\Verizon-Burnsheet\public\India-Rate.xlsx"
SOW_HOURS_FILE = r"C:\Verizon Project\Verizon-Burnsheet\public\SOW_Hours_Rules.xlsx"
US_RATES_FILE = r"C:\Verizon Project\Verizon-Burnsheet\public\Rates and Roles .xlsx"
COMBINED_INPUT_FILE = r"C:\Verizon Project\Verizon-Burnsheet\public\Combined-Input.xlsx"


# ===========================================
# 1. Create IndiaRateCard Dictionary
# ===========================================
def create_india_rate_card(file_path, key_column, value_column, start_row=2):
    """
    Create IndiaRateCard dictionary from Excel file
    
    Args:
        file_path: Path to India-Rate.xlsx
        key_column: Column name to use as dictionary key (e.g., "Concatenate")
        value_column: Column name to use as dictionary value (e.g., "India- ODC / OnPrem")
        start_row: Starting row number (default 2, skipping header)
    
    Returns:
        Dictionary with key-value pairs from Excel
    """
    rate_card = {}
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find column indices by header
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(1, col_idx)
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        
        if key_column not in headers or value_column not in headers:
            print(f"❌ Error: Columns '{key_column}' or '{value_column}' not found in {file_path}")
            return rate_card
        
        key_col_idx = headers[key_column]
        value_col_idx = headers[value_column]
        
        # Iterate through rows and build dictionary
        for row_idx in range(start_row, ws.max_row + 1):
            key_cell = ws.cell(row_idx, key_col_idx)
            value_cell = ws.cell(row_idx, value_col_idx)
            
            if not key_cell.value:  # Skip empty keys
                continue
            
            key = str(key_cell.value).strip()
            value = value_cell.value
            
            # Skip if value is empty
            if value is None:
                continue
            
            # Keep latest row's value for duplicate keys
            rate_card[key] = value
        
        wb.close()
        print(f"✅ IndiaRateCard created with {len(rate_card)} entries")
        return rate_card
    
    except Exception as e:
        print(f"❌ Error reading {file_path}: {str(e)}")
        return rate_card


# ===========================================
# 2. Create hourlyRules Dictionary
# ===========================================
def create_hourly_rules(file_path):
    """
    Create hourlyRules dictionary from SOW_Hours_Rules.xlsx
    
    Args:
        file_path: Path to SOW_Hours_Rules.xlsx
    
    Returns:
        Dictionary mapping Talent Type to Hours Rule
    """
    hourly_rules = {}
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find column indices by header
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(1, col_idx)
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        
        if "Talent Type" not in headers or "Hours Rule" not in headers:
            print(f"❌ Error: Required columns not found in {file_path}")
            return hourly_rules
        
        talent_col_idx = headers["Talent Type"]
        hours_col_idx = headers["Hours Rule"]
        
        # Iterate through rows and build dictionary
        for row_idx in range(2, ws.max_row + 1):
            talent_cell = ws.cell(row_idx, talent_col_idx)
            hours_cell = ws.cell(row_idx, hours_col_idx)
            
            if not talent_cell.value:
                continue
            
            talent = str(talent_cell.value).strip()
            hours = hours_cell.value
            
            if hours is None:
                continue
            
            hourly_rules[talent] = hours
        
        wb.close()
        print(f"✅ Hourly Rules created with {len(hourly_rules)} entries")
        return hourly_rules
    
    except Exception as e:
        print(f"❌ Error reading {file_path}: {str(e)}")
        return hourly_rules


# ===========================================
# 3. Create USRateCard Dictionary
# ===========================================
def create_us_rate_card(file_path):
    """
    Create USRateCard dictionary from Rates and Roles.xlsx
    
    Args:
        file_path: Path to Rates and Roles.xlsx
    
    Returns:
        Dictionary with keys as "ColumnName + RowName" and values as cell values
    """
    rate_card = {}
    
    try:
        wb = load_workbook(file_path)
        
        # Find USA Rate sheet
        usa_sheet = None
        for sheet_name in wb.sheetnames:
            if "USA" in sheet_name.upper() or "US" in sheet_name.upper():
                usa_sheet = wb[sheet_name]
                break
        
        if not usa_sheet:
            print(f"❌ Error: USA Rate sheet not found in {file_path}")
            return rate_card
        
        # Extract column headers from row 2, starting from column B
        headers = {}
        for col_idx in range(2, usa_sheet.max_column + 1):  # Column B = 2
            cell = usa_sheet.cell(2, col_idx)
            if cell.value:
                headers[col_idx] = str(cell.value).strip()
        
        # Extract row labels from column A, starting from row 3
        row_labels = {}
        for row_idx in range(3, usa_sheet.max_row + 1):
            cell = usa_sheet.cell(row_idx, 1)  # Column A = 1
            if cell.value:
                row_labels[row_idx] = str(cell.value).strip()
        
        # Create dictionary with concatenated keys
        for row_idx, row_label in row_labels.items():
            for col_idx, col_header in headers.items():
                key = col_header + row_label  # e.g., "SeniorConsultant"
                value = usa_sheet.cell(row_idx, col_idx).value
                
                if value is not None:
                    rate_card[key] = value
        
        wb.close()
        print(f"✅ USRateCard created with {len(rate_card)} entries")
        return rate_card
    
    except Exception as e:
        print(f"❌ Error reading {file_path}: {str(e)}")
        return rate_card


# ===========================================
# 4. Update Combined-Input.xlsx
# ===========================================
def update_combined_input(india_rate_card, hourly_rules, us_rate_card):
    """
    Update Combined-Input.xlsx with calculated rates
    
    Args:
        india_rate_card: Dictionary with India rates
        hourly_rules: Dictionary with hourly rules
        us_rate_card: Dictionary with US rates
    """
    try:
        wb = load_workbook(COMBINED_INPUT_FILE)
        ws = wb.active
        
        # Find column indices
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(1, col_idx)
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        
        print(f"\n📊 Available columns: {list(headers.keys())}")
        
        # Track updates
        updates_india = 0
        updates_us = 0
        
        # Process data rows
        for row_idx in range(2, ws.max_row + 1):
            # Get Key column value
            if "Key" not in headers:
                continue
            
            key_cell = ws.cell(row_idx, headers["Key"])
            key_value = key_cell.value
            
            if not key_value:
                continue
            
            key_value = str(key_value).strip()
            
            # Get Country column value
            country = None
            if "Country" in headers:
                country_cell = ws.cell(row_idx, headers["Country"])
                country = str(country_cell.value).strip() if country_cell.value else None
            
            # Get Location column value
            location = None
            if "Location" in headers:
                location_cell = ws.cell(row_idx, headers["Location"])
                location = str(location_cell.value).strip() if location_cell.value else None
            
            # === INDIA RATE CALCULATION ===
            if country and country.lower() == "india":
                # Fill Hourly Rate (Rs)
                if key_value in india_rate_card and "Hourly Rate(Rs)" in headers:
                    hourly_rate_rs = india_rate_card[key_value]
                    ws.cell(row_idx, headers["Hourly Rate(Rs)"]).value = hourly_rate_rs
                    updates_india += 1
                
                # Fill Hourly Rate ($) by dividing by 86
                if "Hourly Rate(Rs)" in headers and "Hourly Rate($)" in headers:
                    hourly_rate_rs_cell = ws.cell(row_idx, headers["Hourly Rate(Rs)"])
                    if hourly_rate_rs_cell.value and isinstance(hourly_rate_rs_cell.value, (int, float)):
                        hourly_rate_usd = hourly_rate_rs_cell.value / 86
                        ws.cell(row_idx, headers["Hourly Rate($)"]).value = round(hourly_rate_usd, 2)
                
                # Fill projected/actual rates for ODC or OnPrem
                if location and location in ["ODC", "OnPrem"]:
                    # Fetch India Talent only Key from hourlyRules
                    india_talent_key = None
                    for talent_type, hours_rule in hourly_rules.items():
                        talent_lower = talent_type.lower()
                        if "india" in talent_lower and "only" in talent_lower:
                            india_talent_key = talent_type
                            break
                    
                    if india_talent_key and "Hourly Rate($)" in headers and hourly_rules.get(india_talent_key):
                        hourly_rate_usd_cell = ws.cell(row_idx, headers["Hourly Rate($)"])
                        if hourly_rate_usd_cell.value and isinstance(hourly_rate_usd_cell.value, (int, float)):
                            projected_rate = hourly_rate_usd_cell.value * hourly_rules[india_talent_key]
                            
                            # Fill rate columns
                            rate_columns = ["Projected Rate($)", "Actual Rate", "Jan-26", "Feb-26", "Mar-26"]
                            for col_name in rate_columns:
                                if col_name in headers:
                                    ws.cell(row_idx, headers[col_name]).value = round(projected_rate, 2)
            
            # === USA RATE CALCULATION ===
            else:  # Non-India locations
                # Fill Hourly Rate ($)
                if key_value in us_rate_card and "Hourly Rate($)" in headers:
                    hourly_rate_usd = us_rate_card[key_value]
                    ws.cell(row_idx, headers["Hourly Rate($)"]).value = hourly_rate_usd
                    updates_us += 1
                
                # Fill projected rates based on location
                if location:
                    if location in ["Onshore Verizon Location", "Onshore Supplier Location"]:
                        # Fetch Onshore Talent (USA/Canada) Key value from hourlyRules
                        onshore_key = None
                        for talent_type, hours_rule in hourly_rules.items():
                            talent_lower = talent_type.lower()
                            if "onshore" in talent_lower and ("usa" in talent_lower or "canada" in talent_lower):
                                onshore_key = talent_type
                                break
                        
                        if onshore_key and "Hourly Rate($)" in headers and "Projected Rate($)" in headers and hourly_rules.get(onshore_key):
                            hourly_rate_usd_cell = ws.cell(row_idx, headers["Hourly Rate($)"])
                            if hourly_rate_usd_cell.value and isinstance(hourly_rate_usd_cell.value, (int, float)):
                                projected_rate = hourly_rate_usd_cell.value * hourly_rules[onshore_key]
                                ws.cell(row_idx, headers["Projected Rate($)"]).value = round(projected_rate, 2)
                                
                                # Fill Actual Rate, Jan-26, Feb-26, Mar-26 with same projected rate value
                                rate_columns = ["Actual Rate", "Jan-26", "Feb-26", "Mar-26"]
                                for col_name in rate_columns:
                                    if col_name in headers:
                                        ws.cell(row_idx, headers[col_name]).value = round(projected_rate, 2)
                    
                    elif location.lower() == "india":
                        # Fetch Offshore Talent billed under Onshore Key from hourlyRules
                        offshore_key = None
                        for talent_type, hours_rule in hourly_rules.items():
                            talent_lower = talent_type.lower()
                            if "offshore" in talent_lower and "onshore" in talent_lower:
                                offshore_key = talent_type
                                break
                        
                        if offshore_key and "Hourly Rate($)" in headers and "Projected Rate($)" in headers and hourly_rules.get(offshore_key):
                            hourly_rate_usd_cell = ws.cell(row_idx, headers["Hourly Rate($)"])
                            if hourly_rate_usd_cell.value and isinstance(hourly_rate_usd_cell.value, (int, float)):
                                projected_rate = hourly_rate_usd_cell.value * hourly_rules[offshore_key]
                                ws.cell(row_idx, headers["Projected Rate($)"]).value = round(projected_rate, 2)
                                
                                # Fill Actual Rate, Jan-26, Feb-26, Mar-26 with same projected rate value
                                rate_columns = ["Actual Rate", "Jan-26", "Feb-26", "Mar-26"]
                                for col_name in rate_columns:
                                    if col_name in headers:
                                        ws.cell(row_idx, headers[col_name]).value = round(projected_rate, 2)
        
        # Save the updated workbook
        wb.save(COMBINED_INPUT_FILE)
        wb.close()
        print(f"\n✅ Updates completed!")
        print(f"   India rates filled: {updates_india}")
        print(f"   USA rates filled: {updates_us}")
        print(f"✅ Combined-Input.xlsx saved successfully!")
        
    except Exception as e:
        print(f"❌ Error updating Combined-Input.xlsx: {str(e)}")


def main():
    """Main function to run rate calculation"""
    print("=" * 60)
    print("🚀 Starting Rate Calculation Process")
    print("=" * 60)
    
    # Verify input files exist
    if not os.path.exists(INDIA_RATE_FILE):
        print(f"❌ Error: {INDIA_RATE_FILE} not found")
        return
    
    if not os.path.exists(SOW_HOURS_FILE):
        print(f"❌ Error: {SOW_HOURS_FILE} not found")
        return
    
    if not os.path.exists(US_RATES_FILE):
        print(f"❌ Error: {US_RATES_FILE} not found")
        return
    
    if not os.path.exists(COMBINED_INPUT_FILE):
        print(f"❌ Error: {COMBINED_INPUT_FILE} not found")
        return
    
    print("\n📂 Loading source files...\n")
    
    # Create rate cards
    india_rate_card = create_india_rate_card(
        INDIA_RATE_FILE,
        key_column="Concatenate",
        value_column="India- ODC / OnPrem",
        start_row=2
    )
    
    hourly_rules = create_hourly_rules(SOW_HOURS_FILE)
    
    us_rate_card = create_us_rate_card(US_RATES_FILE)
    
    # Update Combined-Input.xlsx
    print("\n📝 Updating Combined-Input.xlsx...\n")
    update_combined_input(india_rate_card, hourly_rules, us_rate_card)
    
    print("\n" + "=" * 60)
    print("✅ Rate Calculation Process Completed!")
    print("=" * 60)


if __name__ == "__main__":
    main()
