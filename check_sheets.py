import openpyxl

input_file = "public/Home & Marketing, SOR - Burn 2026.xlsx"
wb = openpyxl.load_workbook(input_file)

print(f"Total sheets: {len(wb.sheetnames)}")
print("\nAll sheets (including hidden):")

india_count = 0
usa_count = 0
found_target = False

for name in wb.sheetnames:
    ws = wb[name]
    is_hidden = ws.sheet_state == 'hidden'
    print(f"  {name} - Hidden: {is_hidden}")
    
    # Count India and USA sheets (based on first character)
    if not is_hidden and name.lower() != 'legend':
        if name and name[0].isdigit():
            india_count += 1
            print("    -> India sheet")
        elif name and name[0].isalpha():
            usa_count += 1
            print("    -> USA sheet")
    
    if "1000458649" in str(name):
        print("    ^^^ FOUND!")
        found_target = True

# Check if the sheet exists but is hidden
if "1000458649 - Arun Prakash" in wb.sheetnames:
    print("\nSheet exists but might be hidden")
else:
    print("\nSheet 1000458649 - Arun Prakash NOT FOUND in workbook")

print(f"\n--- Summary ---")
print(f"India Sheets: {india_count}")
print(f"USA Sheets: {usa_count}")
print(f"Target Sheet Found: {found_target}")
