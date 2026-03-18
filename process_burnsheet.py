import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

# Define file paths
BASE_DIR = Path(__file__).parent
PUBLIC_DIR = BASE_DIR / 'public'

RATES_FILE = PUBLIC_DIR / 'Rates and Roles .xlsx'
HOURS_RULES_FILE = PUBLIC_DIR / 'SOW_Hours_Rules.xlsx'
COMBINED_INPUT_FILE = PUBLIC_DIR / 'Combined-Input.xlsx'

# USD to INR conversion rate
USD_TO_INR = 83.48994082840237

def read_rate_card():
    """
    Read the 'New Rate Card India - T&M' sheet from Rates and Roles.xlsx
    and create a dictionary with 'Lookup' as key and 'USD Rates' as value.
    """
    try:
        # Load with data_only=True to get computed values instead of formulas
        wb = openpyxl.load_workbook(RATES_FILE, data_only=True)
        
        # Find the sheet containing "New Rate Card India"
        target_sheet = None
        for sheet_name in wb.sheetnames:
            if 'New Rate Card India' in sheet_name:
                target_sheet = wb[sheet_name]
                print(f"Found sheet: '{sheet_name}'")
                break
        
        if target_sheet is None:
            print("Warning: Could not find 'New Rate Card India' sheet. Using first sheet.")
            target_sheet = wb[wb.sheetnames[0]]
        
        # Get column indices for Lookup and USD Rates
        lookup_col_idx = None
        usd_rates_col_idx = None
        
        for col_idx in range(1, target_sheet.max_column + 1):
            header = target_sheet.cell(row=1, column=col_idx).value
            if header and 'Lookup' in str(header):
                lookup_col_idx = col_idx
            if header and 'USD Rates' in str(header) or 'USD Rate' in str(header):
                usd_rates_col_idx = col_idx
        
        if lookup_col_idx is None or usd_rates_col_idx is None:
            print(f"Error: Could not find 'Lookup' or 'USD Rates' columns")
            print(f"Available columns: {[target_sheet.cell(row=1, column=i).value for i in range(1, target_sheet.max_column + 1)]}")
            return {}
        
        # Create dictionary with Lookup as key and USD Rates as value
        # Handle duplicate keys - keep the latest row's value
        rate_card = {}
        for row_idx in range(2, target_sheet.max_row + 1):
            lookup = target_sheet.cell(row=row_idx, column=lookup_col_idx).value
            usd_rate = target_sheet.cell(row=row_idx, column=usd_rates_col_idx).value
            if lookup and usd_rate is not None:
                # Convert to string and strip whitespace
                lookup_str = str(lookup).strip()
                # Try to convert rate to float
                try:
                    usd_rate_float = float(usd_rate)
                    rate_card[lookup_str] = usd_rate_float
                except (ValueError, TypeError):
                    pass
        
        print("[OK] Rate Card loaded successfully:")
        for key, value in rate_card.items():
            print(f"  {key}: ${value}")
        return rate_card
    except Exception as e:
        print(f"Error reading rate card: {e}")
        import traceback
        traceback.print_exc()
        return {}

def read_hourly_rules():
    """
    Read the SOW_Hours_Rules.xlsx file and create a dictionary
    with 'Talent Type' as key and 'Hours Rule' as value.
    """
    try:
        # Load with data_only=True to get computed values instead of formulas
        wb = openpyxl.load_workbook(HOURS_RULES_FILE, data_only=True)
        ws = wb[wb.sheetnames[0]]  # Get first sheet
        
        # Get column indices for Talent Type and Hours Rule
        talent_type_col_idx = None
        hours_rule_col_idx = None
        
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header and 'Talent Type' in str(header):
                talent_type_col_idx = col_idx
            if header and 'Hours Rule' in str(header):
                hours_rule_col_idx = col_idx
        
        if talent_type_col_idx is None or hours_rule_col_idx is None:
            print(f"Error: Could not find 'Talent Type' or 'Hours Rule' columns")
            print(f"Available columns: {[ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]}")
            return {}
        
        # Create dictionary with Talent Type as key and Hours Rule as value
        # Handle duplicate keys - keep the latest row's value
        hourly_rules = {}
        for row_idx in range(2, ws.max_row + 1):
            talent_type = ws.cell(row=row_idx, column=talent_type_col_idx).value
            hours_rule = ws.cell(row=row_idx, column=hours_rule_col_idx).value
            if talent_type and hours_rule:
                talent_type_str = str(talent_type).strip()
                hours_rule_str = str(hours_rule).strip()
                hourly_rules[talent_type_str] = hours_rule_str
        
        print("[OK] Hourly Rules loaded successfully:")
        for key, value in hourly_rules.items():
            print(f"  {key}: {value} hrs/month")
        return hourly_rules
    except Exception as e:
        print(f"Error reading hourly rules: {e}")
        import traceback
        traceback.print_exc()
        return {}

def rephrase_talent_level(talent_value):
    """
    Replace talent level keywords with appropriate titles.
    - Level1 or Level 1 -> Junior
    - Level2 or Level 2 -> Intermediate
    - Level3, Level 3, Level4, or Level 4 -> Senior
    
    Examples:
    - "PremiumODCLevel 3" -> "PremiumODCSenior"
    - "StandardODCLevel 1" -> "StandardODCJunior"
    - "Level 3 - Premium SkillsOnshore..." -> "Level 3 - Premium SkillsOnshore..." (unchanged - not a valid format)
    """
    talent_str = str(talent_value).strip()
    
    # Match patterns like "PremiumODCLevel 3", "StandardODCLevel 1", etc.
    # Replace Level1/Level 1 with Junior
    if 'Level 1' in talent_str or 'Level1' in talent_str:
        rephrased = talent_str.replace('Level 1', 'Junior').replace('Level1', 'Junior')
    # Replace Level2/Level 2 with Intermediate
    elif 'Level 2' in talent_str or 'Level2' in talent_str:
        rephrased = talent_str.replace('Level 2', 'Intermediate').replace('Level2', 'Intermediate')
    # Replace Level3/Level 3 or Level4/Level 4 with Senior
    elif 'Level 3' in talent_str or 'Level3' in talent_str or 'Level 4' in talent_str or 'Level4' in talent_str:
        rephrased = talent_str.replace('Level 3', 'Senior').replace('Level3', 'Senior') \
                              .replace('Level 4', 'Senior').replace('Level4', 'Senior')
    else:
        rephrased = talent_str
    
    return rephrased

def extract_numeric_hours(hours_rule_str):
    """
    Extract numeric value from hours rule string.
    E.g., "176 hrs/month" -> 176
    """
    try:
        # Try to extract the first number
        import re
        match = re.search(r'(\d+)', str(hours_rule_str))
        if match:
            return int(match.group(1))
    except:
        pass
    return 0

def find_column_by_pattern(column_indices, *patterns):
    """
    Find a column index that matches any of the given patterns.
    Uses case-insensitive matching and supports partial matches.
    
    Args:
        column_indices: Dict mapping column names to indices
        *patterns: One or more patterns to match
    
    Returns:
        Column index if found, None otherwise
    """
    for pattern in patterns:
        pattern_lower = pattern.lower().strip()
        for col_name in column_indices.keys():
            col_name_lower = col_name.lower().strip()
            # Try exact match first
            if col_name_lower == pattern_lower:
                return column_indices[col_name]
            # Try partial match
            if pattern_lower in col_name_lower:
                return column_indices[col_name]
    return None

def process_combined_input(rate_card, hourly_rules):
    """
    Process the Combined Input file:
    1. Read SanitizedKey from first row
    2. Rephrase talent level
    3. Fill in hourly rates and projected rates
    """
    try:
        wb = openpyxl.load_workbook(COMBINED_INPUT_FILE)
        ws = wb[wb.sheetnames[0]]  # Get first sheet
        
        print(f"\nProcessing sheet: '{ws.title}'")
        print(f"Dimensions: {ws.dimensions}")
        
        # Get column headers and find indices
        header_row = 1
        column_indices = {}
        
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col_idx).value
            if header:
                header_str = str(header).strip()
                column_indices[header_str] = col_idx
        
        print(f"\nColumn Headers: {list(column_indices.keys())}")
        
        # Print column mapping report
        print(f"\n[COLUMN MAPPING REPORT]")
        print("=" * 70)
        print(f"Actual columns found in file:")
        for col_name, col_idx in sorted(column_indices.items(), key=lambda x: x[1]):
            print(f"  Column {col_idx}: '{col_name}'")
        
        # Check if required columns exist (with flexible matching)
        required_cols = ['SanitizedKey', 'Hourly Rate ($)', 'Hourly Rate (Rs)', 'Projected Rate($)', 'Actual']
        print(f"\nSearching for required columns:")
        for col in required_cols:
            found_col = find_column_by_pattern(column_indices, col)
            if found_col:
                print(f"  [OK] '{col}' found at column {found_col}")
            else:
                print(f"  [MISSING] '{col}' NOT found (tried fuzzy matching too)")
        print("=" * 70 + "\n")
        
        # Process each data row
        rows_updated = 0
        for row_idx in range(2, ws.max_row + 1):
            print(f"\n--- Processing Row {row_idx} ---")
            
            # Find SanitizedKey column with flexible matching
            sanitized_key_col = find_column_by_pattern(column_indices, 'SanitizedKey')
            
            if sanitized_key_col:
                sanitized_key = ws.cell(row=row_idx, column=sanitized_key_col).value
                
                if sanitized_key:
                    print(f"Original SanitizedKey: {sanitized_key}")
                    
                    # Rephrase the talent level (store in variable X)
                    X = rephrase_talent_level(sanitized_key)
                    print(f"Rephrased value (X): {X}")
                    
                    # Look up hourly rate in rate_card
                    if X in rate_card:
                        hourly_rate_usd = rate_card[X]
                        print(f"Hourly Rate (USD): ${hourly_rate_usd}")
                        
                        # Find and fill Hourly Rate ($) column with flexible matching
                        hr_usd_col = find_column_by_pattern(column_indices, 'Hourly Rate ($)', 'Hourly Rate USD', 'Hourly Rate')
                        if hr_usd_col:
                            ws.cell(row=row_idx, column=hr_usd_col).value = hourly_rate_usd
                            print(f"[OK] Updated Hourly Rate ($) at column {hr_usd_col}")
                            rows_updated += 1
                        else:
                            print(f"[ERROR] Could not find 'Hourly Rate ($)' column")
                        
                        # Calculate and fill Hourly Rate (Rs) column
                        hourly_rate_inr = hourly_rate_usd * USD_TO_INR
                        hr_inr_col = find_column_by_pattern(column_indices, 'Hourly Rate (Rs)', 'Hourly Rate INR')
                        if hr_inr_col:
                            ws.cell(row=row_idx, column=hr_inr_col).value = hourly_rate_inr
                            print(f"[OK] Updated Hourly Rate (Rs) to INR {hourly_rate_inr:.2f} at column {hr_inr_col}")
                        
                        # Look up hours rule and calculate projected rates
                        if X in hourly_rules:
                            hours_rule = hourly_rules[X]
                            hours_numeric = extract_numeric_hours(hours_rule)
                            projected_rate = hourly_rate_usd * hours_numeric
                            
                            proj_col = find_column_by_pattern(column_indices, 'Projected Rate($)', 'Projected Rate')
                            if proj_col:
                                ws.cell(row=row_idx, column=proj_col).value = projected_rate
                                print(f"[OK] Updated Projected Rate($) to ${projected_rate:.2f} at column {proj_col}")
                            
                            actual_col = find_column_by_pattern(column_indices, 'Actual')
                            if actual_col:
                                ws.cell(row=row_idx, column=actual_col).value = projected_rate
                                print(f"[OK] Updated Actual to ${projected_rate:.2f} at column {actual_col}")
                            
                            print(f"Hours Rule: {hours_rule} ({hours_numeric} hours/month)")
                            print(f"Calculation: ${hourly_rate_usd} x {hours_numeric} = ${projected_rate:.2f}")
                        else:
                            print(f"[WARNING] No hours rule found for '{X}' in hourly_rules")
                            print(f"  Available rules: {list(hourly_rules.keys())}")
                    else:
                        print(f"[ERROR] No rate found for '{X}' in rate card")
                        print(f"  Available rates: {list(rate_card.keys())}")
        
        # Save the processed file
        wb.save(str(COMBINED_INPUT_FILE))
        print(f"\n[OK] Processed file saved to: {COMBINED_INPUT_FILE}")
        print(f"[OK] Total rows with updated Hourly Rate ($): {rows_updated}")
        
        # Verify by re-reading
        print("\n" + "=" * 70)
        print("VERIFICATION: Re-reading file to confirm updates")
        print("=" * 70)
        wb_verify = openpyxl.load_workbook(COMBINED_INPUT_FILE)
        ws_verify = wb_verify[wb_verify.sheetnames[0]]
        
        # Find the column again
        verify_indices = {}
        for col_idx in range(1, ws_verify.max_column + 1):
            header = ws_verify.cell(row=1, column=col_idx).value
            if header:
                verify_indices[str(header).strip()] = col_idx
        
        hr_col_verify = find_column_by_pattern(verify_indices, 'Hourly Rate ($)', 'Hourly Rate')
        if hr_col_verify:
            print(f"Sample verification (first 3 data rows):")
            for row in range(2, min(5, ws_verify.max_row + 1)):
                val = ws_verify.cell(row, hr_col_verify).value
                print(f"  Row {row}, Column {hr_col_verify}: {val}")
        
        print("=" * 70 + "\n")
        return True
    
    except Exception as e:
        print(f"Error processing combined input: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main execution function"""
    print("=" * 70)
    print("BURNSHEET PROCESSING SCRIPT")
    print("=" * 70)
    
    # Step 1: Read rate card
    print("\n[1] Reading Rate Card...")
    rate_card = read_rate_card()
    
    if not rate_card:
        print("ERROR: Failed to read rate card. Aborting.")
        return
    
    # Step 2: Read hourly rules
    print("\n[2] Reading Hourly Rules...")
    hourly_rules = read_hourly_rules()
    
    if not hourly_rules:
        print("ERROR: Failed to read hourly rules. Aborting.")
        return
    
    # Step 3: Process combined input
    print("\n[3] Processing Combined Input...")
    success = process_combined_input(rate_card, hourly_rules)
    
    if success:
        print("\n" + "=" * 70)
        print("[SUCCESS] PROCESSING COMPLETE")
        print("=" * 70)
    else:
        print("\n" + "=" * 70)
        print("[FAILED] PROCESSING FAILED")
        print("=" * 70)

if __name__ == '__main__':
    main()
